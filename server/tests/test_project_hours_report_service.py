import unittest
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.database import Base
from app.models import Project, Task, TimeSlot, User
from app.services.project_hours_report_service import build_project_hours_report, export_project_hours_report


class ProjectHoursReportServiceTest(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        self.db = sessionmaker(bind=engine)()
        self.user = User(id=1, username="manager", display_name="项目负责人", role="分析所所长")
        self.project = Project(id=1, code="P-001", name="项目一", manager_id=1)
        parent = Task(
            id=1, project_id=1, name="LCMS方法开发", task_type="manual",
            est_duration_hours=12, status="running", plan_order=1,
        )
        child_a = Task(
            id=2, project_id=1, parent_id=1, name="方案撰写", task_type="manual",
            est_duration_hours=4, status="completed", plan_order=1,
        )
        child_b = Task(
            id=3, project_id=1, parent_id=1, name="方法验证", task_type="instrument",
            est_duration_hours=8, status="running", plan_order=2,
        )
        self.db.add_all([self.user, self.project, parent, child_a, child_b])
        self.db.commit()

    def tearDown(self):
        self.db.close()

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={})
    def test_superseded_slots_do_not_widen_the_planned_window(self, _actual_hours_map):
        """计划开始／结束只看现行时间槽。

        任务被反复重排后会留下大量作废槽。若把它们一起取极值，得到的是历次
        版本的并集——开始取自一个版本、结束取自另一个版本，这个区间从未真实
        存在过，还会用旧版本宽松的结束时间把系统判定掩盖成「正常」。
        """
        from datetime import datetime
        self.db.add_all([
            TimeSlot(id=10, task_id=3, plan_start=datetime(2026, 8, 28, 13, 0),
                     plan_end=datetime(2026, 8, 28, 14, 0), instrument_id=None,
                     status="cancelled", lifecycle_status="superseded"),
            TimeSlot(id=11, task_id=3, plan_start=datetime(2026, 8, 28, 15, 30),
                     plan_end=datetime(2026, 8, 28, 16, 30), instrument_id=None,
                     status="scheduled", lifecycle_status="active"),
            TimeSlot(id=12, task_id=3, plan_start=datetime(2026, 9, 28, 18, 0),
                     plan_end=datetime(2026, 9, 28, 18, 30), instrument_id=None,
                     status="cancelled", lifecycle_status="superseded"),
        ])
        self.db.commit()

        task = {t.task_name: t for t in build_project_hours_report(self.db, self.user).items[0].tasks}["方法验证"]

        self.assertEqual(datetime(2026, 8, 28, 15, 30), task.planned_start)
        self.assertEqual(datetime(2026, 8, 28, 16, 30), task.planned_end)

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={})
    def test_task_with_only_superseded_slots_has_no_planned_window(self, _actual_hours_map):
        """时间槽全部作废的任务当前没有排程，计划时间留空而不是回退到旧版本。"""
        from datetime import datetime
        self.db.add(TimeSlot(
            id=20, task_id=2, plan_start=datetime(2026, 8, 21, 15, 30),
            plan_end=datetime(2026, 8, 31, 16, 0),
            status="cancelled", lifecycle_status="superseded",
        ))
        self.db.commit()

        task = {t.task_name: t for t in build_project_hours_report(self.db, self.user).items[0].tasks}["方案撰写"]

        self.assertIsNone(task.planned_start)
        self.assertIsNone(task.planned_end)

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={})
    def test_night_run_hours_sum_per_task_and_roll_up(self, _actual_hours_map):
        """夜间运行小时数取时间槽自然时长，作废的槽不计，父任务按子任务汇总。"""
        from datetime import datetime
        self.db.add_all([
            # 方法验证的两段夜跑：20:00-次日 8:30 共 12.5h，再加 3h
            TimeSlot(id=1, task_id=3, plan_start=datetime(2026, 9, 1, 20, 0),
                     plan_end=datetime(2026, 9, 2, 8, 30), is_night_run=True,
                     status="completed", lifecycle_status="active"),
            TimeSlot(id=2, task_id=3, plan_start=datetime(2026, 9, 2, 20, 0),
                     plan_end=datetime(2026, 9, 2, 23, 0), is_night_run=True,
                     status="completed", lifecycle_status="active"),
            # 已作废的夜跑不计
            TimeSlot(id=3, task_id=3, plan_start=datetime(2026, 9, 3, 20, 0),
                     plan_end=datetime(2026, 9, 4, 8, 0), is_night_run=True,
                     status="cancelled", lifecycle_status="superseded"),
            # 非夜跑不计
            TimeSlot(id=4, task_id=3, plan_start=datetime(2026, 9, 4, 8, 30),
                     plan_end=datetime(2026, 9, 4, 17, 30), is_night_run=False,
                     status="completed", lifecycle_status="active"),
        ])
        self.db.commit()

        tasks = {t.task_name: t for t in build_project_hours_report(self.db, self.user).items[0].tasks}

        self.assertEqual(15.5, tasks["方法验证"].night_run_hours)
        self.assertEqual(0.0, tasks["方案撰写"].night_run_hours)
        self.assertEqual(15.5, tasks["LCMS方法开发"].night_run_hours)

    @patch("app.services.project_hours_report_service.task_actual_hours_map")
    def test_rolls_up_project_and_parent_actual_hours(self, actual_hours_map):
        actual_hours_map.return_value = {2: 3.5, 3: 6.0}

        report = build_project_hours_report(self.db, self.user)

        self.assertEqual(1, report.project_count)
        self.assertEqual(12, report.planned_hours)
        self.assertEqual(9.5, report.actual_hours)
        self.assertEqual(9.5, report.items[0].tasks[0].actual_hours)
        self.assertEqual(["LCMS方法开发", "方案撰写", "方法验证"], [task.task_name for task in report.items[0].tasks])

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={2: 3.5, 3: 6.0})
    def test_exports_summary_and_task_detail_sheets(self, _actual_hours_map):
        report = build_project_hours_report(self.db, self.user)

        workbook = load_workbook(export_project_hours_report(report))

        self.assertEqual(["项目汇总", "任务明细"], workbook.sheetnames)
        self.assertEqual("P-001", workbook["项目汇总"]["A2"].value)
        self.assertEqual("LCMS方法开发", workbook["任务明细"]["C2"].value)

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={2: 3.5, 3: 6.0})
    def test_filters_projects_by_keyword(self, _actual_hours_map):
        self.assertEqual(1, build_project_hours_report(self.db, self.user, keyword="P-001").project_count)
        self.assertEqual(1, build_project_hours_report(self.db, self.user, keyword="项目负责人").project_count)
        self.assertEqual(0, build_project_hours_report(self.db, self.user, keyword="不存在").project_count)


if __name__ == "__main__":
    unittest.main()
