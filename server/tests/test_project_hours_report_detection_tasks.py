"""检测任务进入项目工时统计报表。

一个检测任务在业务上就相当于一个项目：它本身也是一条 Project 记录，下面固定
挂一条任务，所以在报表里按一行项目展示，并计入合计。
"""

import unittest
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.database import Base
from app.models import Project, Task, User
from app.services.project_hours_report_service import (
    build_project_hours_report,
    export_project_hours_report,
)


class ProjectHoursReportDetectionTaskTest(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        self.db = sessionmaker(bind=engine)()
        self.director = User(id=1, username="director", display_name="分析所所长", role="分析所所长")
        self.operator = User(id=2, username="operator", display_name="张检测", role="技术员")
        self.outsider = User(id=3, username="outsider", display_name="李无关", role="技术员")
        project = Project(id=1, code="P-001", name="项目一", manager_id=1)
        project_task = Task(
            id=1, project_id=1, name="方法开发", task_type="manual",
            est_duration_hours=12, status="running", plan_order=1, assignee_id=2,
        )
        detection = Project(id=2, code="J-2026-001", name="水样重金属检测",
                            project_kind="detection", manager_id=1)
        detection_task = Task(
            id=2, project_id=2, name="水样重金属检测", task_type="instrument",
            est_duration_hours=6, status="completed", plan_order=1, assignee_id=2,
        )
        self.db.add_all([
            self.director, self.operator, self.outsider,
            project, project_task, detection, detection_task,
        ])
        self.db.commit()

    def tearDown(self):
        self.db.close()

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={1: 10.0, 2: 7.5})
    def test_detection_task_is_one_row_and_counts_into_totals(self, _actual_hours_map):
        report = build_project_hours_report(self.db, self.director)

        self.assertEqual(2, report.project_count)
        self.assertEqual(18, report.planned_hours)
        self.assertEqual(17.5, report.actual_hours)
        detection = {item.project_code: item for item in report.items}["J-2026-001"]
        self.assertEqual("detection", detection.project_kind)
        self.assertEqual("水样重金属检测", detection.project_name)
        self.assertEqual(6, detection.planned_hours)
        self.assertEqual(7.5, detection.actual_hours)
        self.assertEqual(1.5, detection.variance_hours)
        self.assertEqual(["水样重金属检测"], [task.task_name for task in detection.tasks])

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={})
    def test_keyword_and_status_filters_apply_to_detection_tasks(self, _actual_hours_map):
        matched = build_project_hours_report(self.db, self.director, keyword="重金属")
        self.assertEqual(["J-2026-001"], [item.project_code for item in matched.items])

        completed = build_project_hours_report(self.db, self.director, statuses={"completed"})
        self.assertEqual(["J-2026-001"], [item.project_code for item in completed.items])

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={})
    def test_detection_visibility_follows_the_detection_module(self, _actual_hours_map):
        """检测任务的可见范围沿用检测任务模块：执行人看得到自己的，无关的人看不到。"""
        own = build_project_hours_report(self.db, self.operator)
        self.assertIn("J-2026-001", [item.project_code for item in own.items])

        other = build_project_hours_report(self.db, self.outsider)
        self.assertEqual([], [item.project_code for item in other.items])

    @patch("app.services.project_hours_report_service.task_actual_hours_map", return_value={})
    def test_export_marks_the_row_kind(self, _actual_hours_map):
        report = build_project_hours_report(self.db, self.director)

        workbook = load_workbook(export_project_hours_report(report))

        summary = {row[1]: row[0] for row in workbook["项目汇总"].iter_rows(min_row=2, values_only=True)}
        self.assertEqual({"P-001": "项目", "J-2026-001": "检测任务"}, summary)
        detail = {row[1]: row[0] for row in workbook["任务明细"].iter_rows(min_row=2, values_only=True)}
        self.assertEqual({"P-001": "项目", "J-2026-001": "检测任务"}, detail)


if __name__ == "__main__":
    unittest.main()
