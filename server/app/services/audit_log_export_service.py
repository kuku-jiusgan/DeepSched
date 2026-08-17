from __future__ import annotations

from io import BytesIO
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ACTION_LABELS = {
    "user_logged_in": "用户登录",
    "project_created": "新增项目",
    "project_updated": "修改项目",
    "project_deleted": "删除项目",
    "schedule_generated": "生成排程",
    "schedule_rescheduled": "重新排程",
    "schedule_insert_confirmed": "确认插单",
    "task_paused": "暂停任务",
    "HTTP POST": "新增/提交",
    "HTTP PUT": "修改",
    "HTTP PATCH": "修改",
    "HTTP DELETE": "删除",
}

TARGET_LABELS = {
    "project": "项目",
    "schedule": "排程",
    "task": "任务",
    "time_slot": "任务排程时间段",
    "approval_gate": "方案签批",
    "user": "系统用户",
    "api_request": "系统操作",
}


def export_audit_logs(records: list[dict]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "操作日志"
    sheet.append(["时间", "操作人", "分类", "操作", "对象", "结果", "摘要", "变更详情"])
    for record in records:
        sheet.append([
            _format_datetime(record.get("created_at")),
            _operator_label(record.get("user_name")),
            record.get("category_label") or "其他操作",
            record.get("action_label") or ACTION_LABELS.get(record.get("action"), record.get("action") or ""),
            record.get("target_display") or _target_text(record),
            "成功" if record.get("result") == "success" else "失败",
            record.get("summary") or "",
            _detail_text(record.get("changes") or record.get("business_detail")),
        ])
    _format_sheet(sheet)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _format_datetime(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if hasattr(value, "strftime") else str(value or "")


def _operator_label(value) -> str:
    if value == "system":
        return "系统自动任务"
    if value == "anonymous":
        return "未登录用户"
    return str(value or "")


def _target_text(record: dict) -> str:
    detail = record.get("detail") or {}
    if detail.get("target_display"):
        return str(detail["target_display"])
    target = TARGET_LABELS.get(record.get("target_type"), record.get("target_type") or "")
    target_id = record.get("target_id")
    return f"{target} #{target_id}" if target_id else target


def _detail_text(detail) -> str:
    if not detail:
        return "-"
    return json.dumps(detail, ensure_ascii=False, separators=(",", ":"), default=str)


def _format_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [20, 16, 16, 18, 36, 10, 60, 80]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for row in sheet.iter_rows(min_row=2):
        row[-2].alignment = Alignment(wrap_text=True, vertical="top")
        row[-1].alignment = Alignment(wrap_text=True, vertical="top")
