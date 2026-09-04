from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.schemas.project_hours_report_schemas import (
    ProjectHoursItemOut,
    ProjectHoursReportOut,
    ProjectHoursTaskOut,
)
from app.services.detection_task_service import list_detection_tasks
from app.services.project_access_service import list_visible_projects
from app.services.project_actual_hours_service import task_actual_hours_map
from app.services.project_status_service import calculate_project_status
from app.models import AuditLog, TaskExecutionSegment, TimeSlot


def build_project_hours_report(
    db,
    user,
    start_date: date | None = None,
    end_date: date | None = None,
    keyword: str | None = None,
    statuses: set[str] | None = None,
) -> ProjectHoursReportOut:
    projects = _filter_projects(
        _reportable_projects(db, user),
        start_date,
        end_date,
        keyword,
        statuses,
    )
    leaf_task_ids = {
        task.id for project in projects for task in project.tasks
        if not task.children and not task.is_external_gate
    }
    leaf_actual_hours = task_actual_hours_map(db, leaf_task_ids)
    task_ids = {task.id for project in projects for task in project.tasks}
    slots_by_task = _slots_by_task(db, task_ids)
    pauses_by_task = _pauses_by_task(db, task_ids)
    delay_reasons_by_task = _delay_reasons_by_task(db, task_ids)
    delay_hours_by_task = _delay_hours_by_task(db, task_ids)
    night_run_hours_by_task = _night_run_hours_by_task(slots_by_task)
    items = [_project_item(project, leaf_actual_hours, slots_by_task, pauses_by_task, delay_reasons_by_task, delay_hours_by_task, night_run_hours_by_task) for project in projects]
    items.sort(key=lambda item: item.project_code)
    return ProjectHoursReportOut(
        generated_at=datetime.now(),
        project_count=len(items),
        planned_hours=round(sum(item.planned_hours for item in items), 2),
        actual_hours=round(sum(item.actual_hours for item in items), 2),
        items=items,
    )


def _reportable_projects(db, user) -> list:
    """报表覆盖的对象：正式项目，加上检测任务。

    一个检测任务在业务上就相当于一个项目——它本身也是一条 Project 记录，
    只是下面固定挂一个任务，所以在报表里同样按一行项目展示、汇总进合计。

    两类记录各有各的可见范围（检测任务允许执行人看自己的），因此分别走各自
    模块的可见性函数，而不是在这里放宽成同一套角色判断。
    """
    return list_visible_projects(db, user) + list_detection_tasks(db, user)


def export_project_hours_report(report: ProjectHoursReportOut) -> BytesIO:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "项目汇总"
    _append_header(summary, ["类型", "项目编号", "项目名称", "客户", "负责人", "项目开始时间", "项目结束时间", "项目状态", "任务数", "预计工时(h)", "实际工时(h)", "工时差异(h)"])
    for item in report.items:
        summary.append([
            _project_kind_label(item.project_kind), item.project_code, item.project_name, item.client_name or "", item.manager_name or "",
            item.start_date, item.end_date, _project_status_label(item.project_status), item.task_count, item.planned_hours, item.actual_hours, item.variance_hours,
        ])
    detail = workbook.create_sheet("任务明细")
    _append_header(detail, ["类型", "项目编号", "项目名称", "任务名称", "层级", "负责人", "仪器编号", "计划开始", "计划结束", "实际开始", "实际完成", "任务状态", "预计工时(h)", "实际工时(h)", "系统判定", "延期小时数", "夜间运行小时数", "暂停次数", "延期/暂停原因"])
    for item in report.items:
        for task in item.tasks:
            detail.append([
                _project_kind_label(item.project_kind), item.project_code, item.project_name, f"{'  ' * task.depth}{task.task_name}", task.depth + 1,
                task.assignee_name or "", "、".join(task.instrument_codes), task.planned_start, task.planned_end,
                task.actual_start, task.actual_end, _task_status_label(task.status), task.planned_hours, task.actual_hours,
                task.schedule_judgement, task.delay_hours, task.night_run_hours, task.pause_count, "；".join(task.pause_reasons),
            ])
    _format_sheet(summary, [10, 16, 24, 20, 16, 20, 20, 14, 10, 14, 14, 14])
    _format_sheet(detail, [10, 16, 24, 32, 10, 16, 16, 18, 18, 18, 18, 14, 14, 14, 12, 10, 14, 10, 30])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _project_item(project, leaf_actual_hours: dict[int, float], slots_by_task: dict, pauses_by_task: dict, delay_reasons_by_task: dict, delay_hours_by_task: dict, night_run_hours_by_task: dict) -> ProjectHoursItemOut:
    tasks = [task for task in project.tasks if not task.is_external_gate]
    task_by_parent: dict[int | None, list] = {}
    for task in tasks:
        task_by_parent.setdefault(task.parent_id, []).append(task)
    for children in task_by_parent.values():
        children.sort(key=lambda task: (task.plan_order, task.id))
    actual_cache: dict[int, float] = {}

    def actual_hours(task) -> float:
        if task.id not in actual_cache:
            children = task_by_parent.get(task.id, [])
            actual_cache[task.id] = sum(actual_hours(child) for child in children) if children else leaf_actual_hours.get(task.id, 0.0)
        return actual_cache[task.id]

    night_run_cache: dict[int, float] = {}

    def night_run_hours(task) -> float:
        if task.id not in night_run_cache:
            children = task_by_parent.get(task.id, [])
            night_run_cache[task.id] = (
                sum(night_run_hours(child) for child in children) if children
                else night_run_hours_by_task.get(task.id, 0.0)
            )
        return night_run_cache[task.id]

    rows: list[ProjectHoursTaskOut] = []

    def append_task(task, depth: int, top_name: str) -> None:
        slots = slots_by_task.get(task.id, [])
        planned_start = min((slot.plan_start for slot in slots), default=None)
        planned_end = max((slot.plan_end for slot in slots), default=None)
        actual_start = min((slot.actual_start for slot in slots if slot.actual_start), default=None)
        last_segment_end = max((slot.actual_end for slot in slots if slot.actual_end), default=None)
        is_completed = task.status in {"done", "completed"}
        actual_end = last_segment_end if is_completed else None
        judgement = _schedule_judgement(planned_end, actual_end, task.status)
        pause_reasons = delay_reasons_by_task.get(task.id, []) + pauses_by_task.get(task.id, [])
        rows.append(ProjectHoursTaskOut(
            task_id=task.id,
            parent_id=task.parent_id,
            task_name=task.name,
            top_level_task_name=top_name,
            assignee_name=task.assignee_name,
            status=task.status,
            depth=depth,
            planned_hours=round(float(task.est_duration_hours or 0), 2),
            actual_hours=round(actual_hours(task), 2),
            instrument_codes=sorted({slot.instrument.code for slot in slots if slot.instrument}),
            planned_start=planned_start, planned_end=planned_end,
            actual_start=actual_start, actual_end=actual_end,
            schedule_judgement=judgement, delay_hours=round(delay_hours_by_task.get(task.id, 0.0), 2),
            night_run_hours=round(night_run_hours(task), 2),
            pause_count=len(pause_reasons), pause_reasons=pause_reasons,
        ))
        for child in task_by_parent.get(task.id, []):
            append_task(child, depth + 1, top_name)

    top_tasks = task_by_parent.get(None, [])
    for task in top_tasks:
        append_task(task, 0, task.name)
    planned = round(sum(float(task.est_duration_hours or 0) for task in top_tasks), 2)
    actual = round(sum(actual_hours(task) for task in top_tasks), 2)
    return ProjectHoursItemOut(
        project_id=project.id,
        project_kind=project.project_kind,
        project_code=project.code,
        project_name=project.name,
        client_name=project.client_name,
        manager_name=project.manager_name,
        start_date=project.start_date,
        end_date=project.end_date,
        project_status=calculate_project_status(project),
        task_count=len(rows),
        planned_hours=planned,
        actual_hours=actual,
        variance_hours=round(actual - planned, 2),
        tasks=rows,
    )


def _night_run_hours_by_task(slots_by_task: dict[int, list]) -> dict[int, float]:
    """每个任务的夜间运行小时数。

    传进来的已经是现行时间槽（_slots_by_task 统一过滤掉了作废的），这里不再
    重复判断，避免两处口径将来各改各的。

    用时间槽的自然时长而不是有效工时：夜跑整段都在工作时段之外，按工作日历
    折算恒为 0，而我们要的正是这段被排除在工时口径之外的仪器占用时长。
    """
    result: dict[int, float] = {}
    for task_id, slots in slots_by_task.items():
        hours = sum(
            (slot.plan_end - slot.plan_start).total_seconds() / 3600
            for slot in slots
            if slot.is_night_run and slot.plan_start and slot.plan_end
        )
        if hours:
            result[task_id] = hours
    return result


def _slots_by_task(db, task_ids: set[int]) -> dict[int, list]:
    """任务的现行时间槽。

    必须排除已作废的槽。报表的计划开始／计划结束是对这批槽取最早开始和最晚
    结束，把历次被推翻的版本混进来，得到的就不是某一版计划，而是所有版本的
    并集——头取自一个版本、尾取自另一个版本，这个区间从未真实存在过，还会
    让系统判定被旧版本宽松的结束时间掩盖成「正常」。仪器编号同理会列出旧版本
    才用到的仪器。甘特图、首页和仪器占用判定用的都是这个条件。
    """
    result: dict[int, list] = {}
    if not task_ids:
        return result
    slots = db.query(TimeSlot).filter(
        TimeSlot.task_id.in_(task_ids),
        TimeSlot.lifecycle_status == "active",
    ).all()
    for slot in slots:
        result.setdefault(slot.task_id, []).append(slot)
    return result


def _pauses_by_task(db, task_ids: set[int]) -> dict[int, list[str]]:
    result: dict[int, list[str]] = {}
    if not task_ids:
        return result
    segments = db.query(TaskExecutionSegment).filter(
        TaskExecutionSegment.task_id.in_(task_ids),
        TaskExecutionSegment.pause_reason.isnot(None),
    ).all()
    for segment in segments:
        reason = (segment.pause_reason or "").strip()
        if reason:
            result.setdefault(segment.task_id, []).append(reason)
    return result


def _delay_reasons_by_task(db, task_ids: set[int]) -> dict[int, list[str]]:
    result: dict[int, list[str]] = {}
    if not task_ids:
        return result
    logs = db.query(AuditLog).filter(AuditLog.action == "task_delay_reported").order_by(AuditLog.created_at.asc()).all()
    for log in logs:
        detail = log.detail if isinstance(log.detail, dict) else {}
        task_id = detail.get("task_id") or (log.target_id if log.target_type == "task" else None)
        reason = str(detail.get("reason") or "").strip()
        if task_id in task_ids and reason and reason not in result.setdefault(task_id, []):
            result[task_id].append(reason)
    return result


def _delay_hours_by_task(db, task_ids: set[int]) -> dict[int, float]:
    result: dict[int, float] = {}
    if not task_ids:
        return result
    logs = db.query(AuditLog).filter(AuditLog.action == "task_delay_reported").all()
    for log in logs:
        detail = log.detail if isinstance(log.detail, dict) else {}
        task_id = detail.get("task_id") or (log.target_id if log.target_type == "task" else None)
        if task_id in task_ids:
            result[task_id] = result.get(task_id, 0.0) + float(detail.get("delay_hours") or 0)
    return result


def _schedule_judgement(planned_end, actual_end, status: str) -> str:
    if not planned_end:
        return ""
    if status in {"done", "completed"}:
        if not actual_end:
            return "正常"
        seconds = (actual_end - planned_end).total_seconds()
        if seconds > 0:
            return "延期"
        if seconds < 0:
            return "提前"
        return "正常"
    overdue_seconds = (datetime.now() - planned_end).total_seconds()
    if overdue_seconds > 0:
        return "延期"
    return "正常"


def _task_status_label(status: str) -> str:
    labels = {
        "pending": "待处理", "ready": "待处理", "scheduled": "待执行",
        "running": "运行中", "paused": "已暂停", "blocked": "已阻塞",
        "interrupted": "已中断", "done": "已完成", "completed": "已完成",
        "waiting_external": "等待外部签批", "waiting_approval": "等待签批",
    }
    return labels.get(status, "未知状态")


PROJECT_KIND_LABELS = {"project": "项目", "detection": "检测任务"}


def _project_kind_label(kind: str) -> str:
    return PROJECT_KIND_LABELS.get(kind, "项目")


PROJECT_STATUS_LABELS = {"pending": "未开始", "active": "进行中", "completed": "已完成"}


def _project_status_label(status: str) -> str:
    return PROJECT_STATUS_LABELS.get(status, "未开始")


def parse_project_statuses(value: str | None) -> set[str]:
    """把逗号分隔的状态筛选值解析成集合，忽略无法识别的值。

    项目状态是按任务实时算出来的，不是数据库列，所以只能在取到项目之后再过滤。
    """
    if not value:
        return set()
    return {
        item.strip() for item in value.split(",")
        if item.strip() in PROJECT_STATUS_LABELS
    }


def _filter_projects(
    projects,
    start_date: date | None,
    end_date: date | None,
    keyword: str | None,
    statuses: set[str] | None = None,
):
    start_at = datetime.combine(start_date, time.min) if start_date else None
    end_at = datetime.combine(end_date, time.max) if end_date else None
    normalized_keyword = (keyword or "").strip().lower()
    wanted = statuses or set()
    return [
        project for project in projects
        if (not wanted or calculate_project_status(project) in wanted)
        and (start_at is None or project.start_date is None or project.start_date >= start_at)
        and (end_at is None or project.start_date is None or project.start_date <= end_at)
        and (
            not normalized_keyword
            or normalized_keyword in project.code.lower()
            or normalized_keyword in project.name.lower()
            or normalized_keyword in (project.client_name or "").lower()
            or normalized_keyword in (project.manager_name or "").lower()
        )
    ]


def _append_header(sheet, values: list[str]) -> None:
    sheet.append(values)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")


def _format_sheet(sheet, widths: list[int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
